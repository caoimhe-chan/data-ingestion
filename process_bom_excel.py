import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import os

def detect_tables_in_sheet(df, min_columns=3, consider_whitespace=True):
    """
    Identify separate tables in a sheet by finding empty row dividers.
    Returns list of DataFrames (each representing one table).
    """
    tables = []
    current_table = []

    for index, row in df.iterrows():
        is_empty = row.isna().all()
        if consider_whitespace:
            is_empty = is_empty or (row.astype(str).str.strip() == '').all()

        if is_empty:
            if len(current_table) >= 1:
                table_df = pd.DataFrame(current_table)
                if table_df.shape[1] >= min_columns:
                    tables.append(table_df)
                current_table = []
        else:
            # Non-empty row - check if meets minimum data requirements
            valid_cells = sum(not pd.isna(val) and str(val).strip() != '' for val in row)
            if valid_cells >= min_columns:
                current_table.append(row.tolist())

    # Add the last table if exists and is valid
    if current_table:
        table_df = pd.DataFrame(current_table)
        if table_df.shape[1] >= min_columns:
            tables.append(table_df)

    return tables

def process_table(table_df):
    """Process a single table DataFrame to detect headers and clean data."""
    # Find header row (first row with most non-empty values)
    header_candidates = []
    for i, row in table_df.iterrows():
        non_empty = sum(not pd.isna(val) and str(val).strip() != '' for val in row)
        header_candidates.append((i, non_empty))

    if not header_candidates:
        return None
    header_row, _ = max(header_candidates, key=lambda x: x[1])

    headers = table_df.iloc[header_row].values
    cleaned_df = table_df.iloc[header_row+1:].copy()
    cleaned_df.columns = headers

    cleaned_df = cleaned_df.dropna(how='all').dropna(axis=1, how='all')
    return cleaned_df.reset_index(drop=True)

def is_stock_number_column(column_data, column_name):
    """Check if column is 'Stock Number' based on data + header name."""
    sample = column_data.dropna().astype(str).head(20)
    column_name = str(column_name).lower()

    # Rule 1: Header clearly indicates stock number
    stock_keywords = ["stock", "stock no", "stock id", "stock number", "part no", "part number"]
    if any(keyword in column_name for keyword in stock_keywords):
        return True

    # Rule 2: Data matches stock number patterns
    stock_pattern = r'^[A-Za-z0-9\-_\.]+$'
    if len(sample) > 0 and all(sample.str.match(stock_pattern, na=False)):
        # Exclude pure numbers that might be quantities
        if not all(sample.str.isdigit()):
            return True
        else:
            # Only treat as stock number if long numeric code
            return sample.str.len().mean() > 5
    return False

def detect_column_type(column_data, column_name):
    """Detect column type using both data and header name."""
    column_name_str = str(column_name).lower()
    sample = column_data.dropna().head(20)

    # Skip weight percentage columns
    if 'weight%' in column_name_str:
        return None

    # Priority 1: Check header name first
    if "stock" in column_name_str:
        return "Stock Number"
    if "weight" in column_name_str or any(unit in column_name_str for unit in ["kg", "g", "lb"]):
        return "Weight(kg)"
    if "description" in column_name_str:
        return "Description"
    if "supplier" in column_name_str or "vendor" in column_name_str:
        return "Supplier"
    if "qty" in column_name_str or "quantity" in column_name_str:
        return "Quantity"
    if "material" in column_name_str or "mat" in column_name_str:
        return "Material"

    # Priority 2: Infer from data patterns
    if is_stock_number_column(column_data, column_name):
        return "Stock Number"
    if len(sample) > 0 and pd.api.types.is_numeric_dtype(sample):
        return "Quantity"

    # Default to Notes
    return "Notes"

def extract_number_from_text(text):
    """Extract numeric value from text, handling various formats."""
    if pd.isna(text):
        return 0.0

    text_str = str(text).strip()
    if text_str == '' or text_str.lower() in ['nan', 'none']:
        return 0.0

    # Try direct conversion first
    try:
        return float(text_str)
    except ValueError:
        pass

    # Extract numbers from text using regex
    numbers = re.findall(r'\d+\.?\d*', text_str)
    if numbers:
        return float(numbers[0])

    return 0.0

def clean_and_standardize_bom(df):
    """Standardize the BOM with proper column mapping and weight conversion."""
    STANDARD_COLS = [
        "Stock Number",
        "Description",
        "Material",
        "Quantity",
        "Weight(kg)",
        "Supplier",
        "Notes"
    ]
    try:
        print(f"Input DataFrame shape: {df.shape}")
        print(f"Input columns: {list(df.columns)}")

        # Create mapping from original columns to standard columns
        column_mapping = {}
        used_columns = set()

        # First pass: map columns based on detection
        for col in df.columns:
            detected_type = detect_column_type(df[col], col)
            if detected_type and detected_type not in used_columns:
                column_mapping[col] = detected_type
                used_columns.add(detected_type)
                print(f"Mapped '{col}' -> '{detected_type}'")

        # Create standardized DataFrame
        standardized_df = pd.DataFrame()

        # Map existing columns
        for original_col, standard_col in column_mapping.items():
            standardized_df[standard_col] = df[original_col].copy()

        # Add missing standard columns with default values
        for standard_col in STANDARD_COLS:
            if standard_col not in standardized_df.columns:
                if standard_col == "Weight(kg)":
                    standardized_df[standard_col] = 0.0
                elif standard_col == "Quantity":
                    standardized_df[standard_col] = 1
                else:
                    standardized_df[standard_col] = ""
                print(f"Added missing column '{standard_col}' with default values")

        # Enhanced Weight(kg) conversion
        if "Weight(kg)" in standardized_df.columns:
            print("\nProcessing Weight(kg) column...")
            print(f"Weight column before conversion: {standardized_df['Weight(kg)'].head().tolist()}")

            # Apply number extraction to each value
            standardized_df["Weight(kg)"] = standardized_df["Weight(kg)"].apply(extract_number_from_text)

            print(f"Weight column after conversion: {standardized_df['Weight(kg)'].head().tolist()}")
            print(f"Non-zero weights: {(standardized_df['Weight(kg)'] > 0).sum()}")

        # Clean string columns
        string_columns = ["Stock Number", "Description", "Material", "Supplier", "Notes"]
        for col in string_columns:
            if col in standardized_df.columns:
                standardized_df[col] = standardized_df[col].astype(str).str.strip()
                standardized_df[col] = standardized_df[col].replace('nan', '')

        # Ensure Quantity is numeric
        if "Quantity" in standardized_df.columns:
            standardized_df["Quantity"] = pd.to_numeric(standardized_df["Quantity"], errors='coerce').fillna(1)

        # Remove completely empty rows
        standardized_df = standardized_df.dropna(how='all')

        # Reorder columns to match standard order
        standardized_df = standardized_df[STANDARD_COLS]

        print(f"Output DataFrame shape: {standardized_df.shape}")
        return standardized_df

    except Exception as e:
        print(f"🚨 Error in clean_and_standardize_bom: {e}")
        import traceback
        traceback.print_exc()
        # Return empty DataFrame with standard columns as fallback
        return pd.DataFrame(columns=STANDARD_COLS)

def process_and_export_bom(input_path, output_path, specific_sheets=None, combine_all_sheets=True):
    """
    Process multiple sheets and export to formatted Excel.

    Parameters:
    - input_path: Path to input Excel file
    - output_path: Path for output Excel file
    - specific_sheets: List of sheet names to process (None = all sheets)
    - combine_all_sheets: If True, combine all sheets into one output. If False, create separate output sheets.
    """
    try:
        # Read all sheets or specific sheets
        if specific_sheets:
            sheets = {}
            for sheet_name in specific_sheets:
                try:
                    sheets[sheet_name] = pd.read_excel(input_path, sheet_name=sheet_name)
                    print(f" Loaded sheet: {sheet_name}")
                except Exception as e:
                    print(f" Failed to load sheet '{sheet_name}': {e}")
        else:
            sheets = pd.read_excel(input_path, sheet_name=None)
            print(f" Loaded {len(sheets)} sheets: {list(sheets.keys())}")

    except Exception as e:
        print(f" Error reading Excel file: {e}")
        return False

    all_dfs = []
    sheet_results = {}

    for sheet_name, df in sheets.items():
        print(f"\n{'='*60}")
        print(f"Processing sheet: {sheet_name}")
        print(f"{'='*60}")
        try:
            # Detect and process all tables in sheet
            tables = detect_tables_in_sheet(df)
            if not tables:
                print(f"⚠️ No tables found in sheet '{sheet_name}'")
                continue

            print(f"Found {len(tables)} table(s) in sheet")

            # Process each table and combine
            processed_tables = []
            for i, table in enumerate(tables):
                print(f"\nProcessing table {i+1}...")
                processed = process_table(table)
                if processed is not None and not processed.empty:
                    print(f"Table {i+1} processed: {processed.shape}")
                    standardized = clean_and_standardize_bom(processed)
                    if not standardized.empty:
                        processed_tables.append(standardized)
                        print(f"Table {i+1} standardized: {standardized.shape}")
                    else:
                        print(f"Table {i+1} resulted in empty standardized data")
                else:
                    print(f"Table {i+1} could not be processed")

            if processed_tables:
                combined_table = pd.concat(processed_tables, ignore_index=True)
                combined_table["Source Sheet"] = sheet_name

                # Store results for this sheet
                sheet_results[sheet_name] = combined_table

                if combine_all_sheets:
                    all_dfs.append(combined_table)

                print(f" Sheet '{sheet_name}': Combined {len(combined_table)} rows from {len(processed_tables)} tables")
            else:
                print(f" Sheet '{sheet_name}': No valid tables processed")

        except Exception as e:
            print(f" Error processing sheet '{sheet_name}': {str(e)}")
            import traceback
            traceback.print_exc()

    # Export results
    if combine_all_sheets and all_dfs:
        # Combine all sheets into one output sheet
        final_bom = pd.concat(all_dfs, ignore_index=True)

        print(f"\n{'='*60}")
        print(f"COMBINED RESULTS FROM ALL SHEETS")
        print(f"{'='*60}")
        print(f"Total rows: {len(final_bom)}")
        print(f"Sheets processed: {len(all_dfs)}")
        print(f"Columns: {list(final_bom.columns)}")
        print(f"Weight statistics:")
        print(f"  - Non-zero weights: {(final_bom['Weight(kg)'] > 0).sum()}")
        print(f"  - Average weight: {final_bom['Weight(kg)'].mean():.2f}")
        print(f"  - Weight range: {final_bom['Weight(kg)'].min():.2f} - {final_bom['Weight(kg)'].max():.2f}")

        # Save combined results
        save_excel_with_formatting(final_bom, output_path, "Combined_BOM")

    elif not combine_all_sheets and sheet_results:
        # Create separate sheets for each processed sheet
        print(f"\n{'='*60}")
        print(f"SEPARATE RESULTS FOR EACH SHEET")
        print(f"{'='*60}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in sheet_results.items():
                safe_sheet_name = sheet_name[:31]  # Excel sheet name limit
                df_to_save = df.drop(columns=['Source Sheet'], errors='ignore')

                df_to_save.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                apply_excel_formatting(writer.book, writer.sheets[safe_sheet_name])

                print(f" Sheet '{safe_sheet_name}': {len(df)} rows")

        print(f"\n Success! Separate sheets saved to: {output_path}")

    else:
        print(" No valid data processed from any sheet")
        return False

    return True

def save_excel_with_formatting(df, output_path, sheet_name="Standardized BOM"):
    """Save DataFrame to Excel with formatting."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        apply_excel_formatting(writer.book, writer.sheets[sheet_name])

    print(f"\n Success! Standardized BOM saved to: {output_path}")

def apply_excel_formatting(workbook, worksheet):
    """Apply consistent formatting to Excel worksheet."""
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust columns
    for col in worksheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

def get_sheet_summary(input_path):
    """Get a summary of all sheets in the Excel file."""
    try:
        sheets = pd.read_excel(input_path, sheet_name=None)
        print(f"\n{'='*60}")
        print(f"EXCEL FILE SUMMARY")
        print(f"{'='*60}")
        print(f"File: {input_path}")
        print(f"Total sheets: {len(sheets)}")

        for i, (sheet_name, df) in enumerate(sheets.items(), 1):
            print(f"\n{i}. Sheet: '{sheet_name}'")
            print(f"   Shape: {df.shape}")
            print(f"   Columns: {list(df.columns)[:5]}{'...' if len(df.columns) > 5 else ''}")

        return list(sheets.keys())
    except Exception as e:
        print(f" Error reading file: {e}")
        return []

def main():
    file_path = input("Enter the path to the Excel file: ")

    if not os.path.exists('Data'):
        os.makedirs('Data')

    output_path = os.path.join('Data', 'processed_bom.xlsx')

    print("\nSelect export mode:")
    print("1. Combine all sheets into one output sheet.")
    print("2. Create a separate output sheet for each original sheet.")

    while True:
        try:
            choice = int(input("Enter your choice (1 or 2): "))
            if choice in [1, 2]:
                combine_all_sheets = (choice == 1)
                break
            else:
                print("Invalid choice. Please enter 1 or 2.")
        except ValueError:
            print("Invalid input. Please enter a number.")

    process_and_export_bom(
        input_path=file_path,
        output_path=output_path,
        combine_all_sheets=combine_all_sheets
    )


if __name__ == '__main__':
    main()
